import json
import os
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side
from openpyxl.styles.numbers import NumberFormat
from requests import get, post
from shutil import unpack_archive
from time import sleep

MAX_RETRIES = 5
TICKS = 50
TIME_090000 = {
    "start" : 90000, 
    "end" : 90300
}
TIME_133000 = {
    "start" : 133000,
    "end" : 133300
}

def find_latest_date_in_excel(sheet) -> datetime:
    latest_date = None
    for cell in sheet["G"]: # 使用 Call 成交價進行確認
        if cell.value:
            latest_date = sheet[f"A{cell.row}"].value
        else:
            write_row = cell.row
            break
    
    if not isinstance(latest_date, datetime):
        latest_date = datetime.strptime(latest_date, "%Y/%m/%d")
    
    return latest_date, write_row


def get_previous_30_trading_days() -> list:
    url = f"https://www.taifex.com.tw/cht/3/optPrevious30DaysSalesData"
    print(f"開始取得前 30 個交易日期 {url}...")
    for retry in range(MAX_RETRIES):
        try:
            res = get(url, timeout=10)
            res.raise_for_status()

            dates_30 = []
            soup = BeautifulSoup(res.text, "html.parser")
            datas = soup.select("table.table_f td:nth-of-type(2)")
            for data in datas:
                dates_30.append(data.text)
            print(f"{dates_30=}")
            return dates_30

        except Exception as e:
            print(f"重新取得中 ({e})... {retry+1}/{MAX_RETRIES}")
            sleep(3)
    else:
        print("資料取得失敗")
        os._exit(1)


def get_option_daily_zip(zip_filename):
    url = f"https://www.taifex.com.tw/file/taifex/Dailydownload/OptionsDailydownloadCSV/{zip_filename}"
    print(f"開始下載 {url} ...")
    for retry in range(MAX_RETRIES):
        try:
            res = get(url, timeout=10)
            res.raise_for_status()
            with open(f"Options/{zip_filename}", 'wb') as f:
                f.write(res.content)
            print(f"{zip_filename} 下載成功 !")
            return
        except Exception as e:
            print(f"重新下載中 ({e})... {retry+1}/{MAX_RETRIES}")
            sleep(3)
    else:
        print("下載資料失敗")
        os._exit(1)


def get_twse_open_close(dates):
    current_month = None
    datas = {}
    for _d in dates:
        _date = datetime.strptime(_d, "%Y/%m/%d")
        _month = _date.month
        if current_month == _month:
            continue
        current_month = _month
        first_day_of_month = _date.replace(day=1).strftime("%Y%m%d")
        url = f"https://www.twse.com.tw/rwd/zh/TAIEX/MI_5MINS_HIST?date={first_day_of_month}&response=json"
        print(f"開始取得前 {current_month} 月，加權指數價格 {url} ...")
        for retry in range(MAX_RETRIES):
            try:
                res = get(url, timeout=10)
                res.raise_for_status()
                res = json.loads(res.text)

                for row in res["data"]:
                    # row = ['113/07/01', '23,042.70', '23,187.88', '23,015.17', '23,058.57']
                    bc_date = row[0].split("/")
                    bc_date[0] = str(int(bc_date[0]) + 1911)
                    bc_date = "/".join(bc_date)

                    _open = row[1].replace(",","").split(".")[0]
                    _close = row[4].replace(",","").split(".")[0]

                    datas[bc_date] = {"open": _open, "close": _close}
                break

            except Exception as e:
                print(f"重新取得中 ({e})... {retry+1}/{MAX_RETRIES}")
                sleep(3)
        else:
            print("資料取得失敗")
            os._exit(1)
        sleep(3)
    return datas


def get_week_code(date, for_1330=True):
    current_weekday = date.weekday()  # 星期幾（0=星期一, 6=星期日）
    delta_day_to_Wednesday = (2-current_weekday + 7) % 7 # 星期三為 2
    next_Wednesday = date + timedelta(days=delta_day_to_Wednesday)

    if next_Wednesday.day % 7 == 0: # 判斷是當月第幾個禮拜三
        week_num = next_Wednesday.day // 7
    else:
        week_num = next_Wednesday.day // 7 + 1

    if week_num != 3:
        week_code_for_0900 = f"{next_Wednesday.year}{next_Wednesday.month:02d}W{week_num}"
    else:
        week_code_for_0900 = f"{next_Wednesday.year}{next_Wednesday.month:02d}"

    if for_1330:
        if current_weekday == 2: # 如果是結算日則 1330 需要看下一期的
            next_Wednesday = date + timedelta(days=7)
            week_code_for_1330 = get_week_code(next_Wednesday, for_1330=False)
        else:
            week_code_for_1330 = week_code_for_0900
        return week_code_for_0900, week_code_for_1330
    else:
        return week_code_for_0900


def get_call_and_put(specific_date, _type, week_code, csv_filename, twse_dict, search_range):
    df = pd.read_csv(csv_filename, encoding="big5", low_memory=False)

    df.columns = df.columns.str.strip()  # 去除欄位名稱的空格
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)  # 去除每個欄位值的空格

    twse_point = int(twse_dict[specific_date][_type])
    print(f"{specific_date} 大盤 {_type} 為: {twse_point}" )
    strike_lists = []
    remains = twse_point % TICKS

    for i in range(0, 6):
        strike_lists.append(twse_point + (TICKS - remains) + TICKS * i)
        strike_lists.append(twse_point - remains - TICKS * i)
    if remains == 0:
        strike_lists.append(twse_point)

    strike_lists.sort()

    print(f"取的 {strike_lists} 的 Call 與 Put 數值")

    last_strike, last_call_deal, last_put_deal = 0, 0, 0
    call_strike, call_deal, put_strike, put_deal = 0, 0, 0, 0

    for current_strike in strike_lists:
        
        filtered_df = df[(df['商品代號'] == 'TXO') & (df['履約價格'] == current_strike) & (df['到期月份(週別)'] == week_code) & (df['成交時間'] >= search_range["start"]) & (df['成交時間'] <= search_range["end"])]
        
        call_min = filtered_df[filtered_df['買賣權別'] == 'C']["成交價格"].min()
        put_max = filtered_df[filtered_df['買賣權別'] == 'P']["成交價格"].max()

        if last_call_deal > last_put_deal and put_max > call_min:
            call_strike = current_strike
            call_deal = call_min
            put_strike = last_strike
            put_deal = last_put_deal
            break
        else:
            last_strike = current_strike
            last_call_deal = call_min if str(call_min) != "nan" else last_call_deal
            last_put_deal = put_max if str(put_max) != "nan" else last_put_deal

    return call_strike, call_deal, put_strike, put_deal


def get_call_and_put_special(df, week_code, final_time, Call_dict, Put_dict):
    """
    如果基本方式找不到,則使用以下
    抓取當下時間的最小與最大履約金,
    並依照每個 tick 迴圈,
    將每個履約價的 C 與 P 補齊，
    補齊方法為, 從 final_time["start"] 開始判斷有無
    """
    min_strike_price = min(list(Call_dict) + list(Put_dict))
    max_strike_price = max(list(Call_dict) + list(Put_dict))
    print(f"{min_strike_price=}")
    print(f"{max_strike_price=}")

    for _sp in range(min_strike_price, max_strike_price+1, TICKS):
        if _sp not in Call_dict: 
            search_time = final_time["start"] # 從一開始找
            print(f"補 {_sp} 到 Call ===================================")
            while search_time <= final_time["end"]:
                search_time_int = int(search_time.strftime('%H%M%S'))
                # print(f"當前查找時間 = {search_time_int}")

                filtered_df = df[(df['商品代號'] == 'TXO') & (df['買賣權別'] == 'C') & (df['成交時間'] == search_time_int) & (df['到期月份(週別)'] == week_code) & (df['履約價格'] == int(_sp))]
                if not filtered_df.empty:
                    final_price = filtered_df.loc[filtered_df['成交價格'].idxmin()]['成交價格']
                    # print(filtered_df.to_dict())
                    # print(final_price)
                    Call_dict[int(_sp)] = final_price
                    break
                
                search_time += timedelta(seconds=1)
            else:
                # 會走到這邊代表這一分鐘內 根本沒有 這個履約價，直接給予 0
                return False

        # if _sp not in Put_dict: 
        #     search_time = final_time["start"] # 從一開始找
        #     print(f"補 {_sp} 到 Put ===================================")
        #     while search_time <= final_time["end"]:
        #         search_time_int = int(search_time.strftime('%H%M%S'))
        #         # print(f"當前查找時間 = {search_time_int}")

        #         filtered_df = df[(df['商品代號'] == 'TXO') & (df['買賣權別'] == 'P') & (df['成交時間'] == search_time_int) & (df['到期月份(週別)'] == week_code) & (df['履約價格'] == int(_sp))]
        #         if not filtered_df.empty:
        #             final_price = filtered_df.loc[filtered_df['成交價格'].idxmax()]['成交價格']
        #             # print(filtered_df.to_dict())
        #             # print(final_price)
        #             Put_dict[int(_sp)] = final_price
        #             break
                
        #         search_time += timedelta(seconds=1)
        #     else:
        #         return False

    print(Call_dict)
    # return Call_dict, Put_dict
    os._exit(1)


def get_settlement_price(date):

    date_str = date.strftime("%Y/%m/%d")

    url = "https://www.taifex.com.tw/cht/5/optIndxFSP"
    payloads = {
        "commodityIds": "2",
        "_all": "on",
        "start_year": str(date.year),
        "start_month": f"{date.month:02d}",
        "end_year": str(date.year),
        "end_month": f"{date.month:02d}",
        "button": "送出查詢",
    }
    print(f"開始取得 {date_str} 結算價 - {url} ...")
    for retry in range(MAX_RETRIES):
        try:
            res = post(url, data=payloads, timeout=10)
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "html.parser")
            datas = soup.select("table.table_f tbody tr")
            for data in datas:
                settlement_date = data.find_all("td")[0].text.strip()
                if date_str == settlement_date: # 確認是相同日期，則取往下得當週結算價
                    settlement_price = int(data.find_all("td")[2].text)
                    print(f"{date_str} 結算價為 {settlement_price}")
                    return settlement_price
            raise
 
        except Exception as e:
            print(f"重新取得 {date} 結算價中 ({e})... {retry+1}/{MAX_RETRIES}")
            sleep(3)
    else:
        print(f"取得 {date} 結算價失敗")
        os._exit(1)


if __name__ == "__main__":

    os.makedirs("Options", exist_ok=True)
    wb = load_workbook("Options.xlsx")
    try:
        wb.save("Options.xlsx")
    except:
        print("關閉 Options.xlsx 後, 再次執行程式")
        os.system("pause")
        os._exit(1)
    sheet = wb["純紀錄"]

    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(name='Microsoft JhengHei UI', size=10)

    excel_latest_date, write_row = find_latest_date_in_excel(sheet)
    print(f"Excel 最新日期為 {excel_latest_date}")

    dates_30 = get_previous_30_trading_days()

    if excel_latest_date.strftime("%Y/%m/%d") not in dates_30:
        print(f"Excel 最新日期為 {excel_latest_date}, 已超過 30 天內能取得之資料")
        os._exit(1)

    twse_dict = get_twse_open_close(dates_30)
    #twse_dict = {'2024/08/01': {'open': '22546', 'close': '22642'}, '2024/08/02': {'open': '22141', 'close': '21638'}, '2024/07/01': {'open': '23042', 'close': '23058'}, '2024/07/02': {'open': '23012', 'close': '22879'}, '2024/07/03': {'open': '23009', 'close': '23172'}, '2024/07/04': {'open': '23360', 'close': '23522'}, '2024/07/05': {'open': '23532', 'close': '23556'}, '2024/07/08': {'open': '23550', 'close': '23878'}, '2024/07/09': {'open': '23888', 'close': '23900'}, '2024/07/10': {'open': '23744', 'close': '24007'}, '2024/07/11': {'open': '24242', 'close': '24390'}, '2024/07/12': {'open': '23955', 'close': '23916'}, '2024/07/15': {'open': '23927', 'close': '23879'}, '2024/07/16': {'open': '23880', 'close': '23997'}, '2024/07/17': {'open': '23827', 'close': '23769'}, '2024/07/18': {'open': '23373', 'close': '23398'}, '2024/07/19': {'open': '23228', 'close': '22869'}, '2024/07/22': {'open': '22818', 'close': '22256'}, '2024/07/23': {'open': '22514', 'close': '22871'}, '2024/07/26': {'open': '22206', 'close': '22119'}, '2024/07/29': {'open': '22321', 'close': '22164'}, '2024/07/30': {'open': '22040', 'close': '22223'}, '2024/07/31': {'open': '22088', 'close': '22199'}, '2024/06/03': {'open': '21388', 'close': '21536'}, '2024/06/04': {'open': '21513', 'close': '21356'}, '2024/06/05': {'open': '21385', 'close': '21484'}, '2024/06/06': {'open': '21856', 'close': '21902'}, '2024/06/07': {'open': '21823', 'close': '21858'}, '2024/06/11': {'open': '21984', 'close': '21792'}, '2024/06/12': {'open': '21841', 'close': '22048'}, '2024/06/13': {'open': '22217', 'close': '22312'}, '2024/06/14': {'open': '22311', 'close': '22504'}, '2024/06/17': {'open': '22468', 'close': '22496'}, '2024/06/18': {'open': '22690', 'close': '22757'}, '2024/06/19': {'open': '22858', 'close': '23209'}, '2024/06/20': {'open': '23196', 'close': '23406'}, '2024/06/21': {'open': '23193', 'close': '23253'}, '2024/06/24': {'open': '23124', 'close': '22813'}, '2024/06/25': {'open': '22695', 'close': '22875'}, '2024/06/26': {'open': '22938', 'close': '22986'}, '2024/06/27': {'open': '22859', 'close': '22905'}, '2024/06/28': {'open': '22896', 'close': '23032'}}

    for _d in dates_30[::-1]:
        current_date = datetime.strptime(_d, "%Y/%m/%d")
        if current_date <= excel_latest_date:
            continue
        current_date_str = current_date.strftime("%Y_%m_%d")
        print(f"執行日期: {current_date_str} ".ljust(80, "="))

        zip_filename = f"OptionsDaily_{current_date_str}.zip"
        get_option_daily_zip(zip_filename)
        unpack_archive(f"Options/{zip_filename}", "Options") # 解壓縮

        week_code_for_0900, week_code_for_1330 = get_week_code(current_date)

        csv_filename = f"Options/OptionsDaily_{current_date_str}.csv"
        call_strike_0900, call_deal_0900, put_strike_0900, put_deal_0900 = get_call_and_put(_d, "open", week_code_for_0900, csv_filename, twse_dict, TIME_090000)
        call_strike_1330, call_deal_1330, put_strike_1330, put_deal_1330 = get_call_and_put(_d, "close", week_code_for_1330, csv_filename, twse_dict, TIME_133000)

        print(f"{put_strike_0900=}")
        print(f"{put_deal_0900=}")
        print(f"{call_strike_0900=}")
        print(f"{call_deal_0900=}")
        print(f"{put_strike_1330=}")
        print(f"{put_deal_1330=}")
        print(f"{call_strike_1330=}")
        print(f"{call_deal_1330=}")

        sheet[f"A{write_row}"].value = current_date
        sheet[f"A{write_row}"].font = Font(name='微軟正黑體', size=8)
        sheet[f"A{write_row}"].number_format = "YYYY/M/D"

        sheet[f"A{write_row+1}"].value = current_date
        sheet[f"A{write_row+1}"].font = Font(name='微軟正黑體', size=8)
        sheet[f"A{write_row+1}"].number_format = "YYYY/M/D"

        sheet[f"E{write_row}"].value = week_code_for_0900
        sheet[f"E{write_row}"].alignment = Alignment(horizontal='right', vertical='center')
        sheet[f"E{write_row}"].font = Font(name='Microsoft JhengHei UI', size=8)
        sheet[f"E{write_row}"].number_format = "General"

        sheet[f"E{write_row+1}"].value = week_code_for_1330
        sheet[f"E{write_row+1}"].alignment = Alignment(horizontal='right', vertical='center')
        sheet[f"E{write_row+1}"].font = Font(name='Microsoft JhengHei UI', size=8)
        sheet[f"E{write_row+1}"].number_format = "General"

        sheet[f"G{write_row}"].value = call_deal_0900
        sheet[f"G{write_row}"].alignment = alignment
        sheet[f"G{write_row}"].font = font
        sheet[f"G{write_row}"].number_format = "0.0"
        
        sheet[f"G{write_row+1}"].value = call_deal_1330
        sheet[f"G{write_row+1}"].alignment = alignment
        sheet[f"G{write_row+1}"].font = font
        sheet[f"G{write_row+1}"].number_format = "0.0"

        sheet[f"H{write_row}"].value = call_strike_0900
        sheet[f"H{write_row}"].alignment = alignment
        sheet[f"H{write_row}"].font = font
        sheet[f"H{write_row}"].number_format = "General"

        sheet[f"H{write_row+1}"].value = call_strike_1330
        sheet[f"H{write_row+1}"].alignment = alignment
        sheet[f"H{write_row+1}"].font = font
        sheet[f"H{write_row+1}"].number_format = "General"

        sheet[f"I{write_row}"].value = put_strike_0900
        sheet[f"I{write_row}"].alignment = alignment
        sheet[f"I{write_row}"].font = font
        sheet[f"I{write_row}"].number_format = "General"

        sheet[f"I{write_row+1}"].value = put_strike_1330
        sheet[f"I{write_row+1}"].alignment = alignment
        sheet[f"I{write_row+1}"].font = font
        sheet[f"I{write_row+1}"].number_format = "General"

        sheet[f"J{write_row}"].value = put_deal_0900
        sheet[f"J{write_row}"].alignment = alignment
        sheet[f"J{write_row}"].font = font
        sheet[f"J{write_row}"].number_format = "0.0"

        sheet[f"J{write_row+1}"].value = put_deal_1330
        sheet[f"J{write_row+1}"].alignment = alignment
        sheet[f"J{write_row+1}"].font = font
        sheet[f"J{write_row+1}"].number_format = "0.0"

        wb.save("Options.xlsx")

        if current_date.weekday() == 2: # 如果是星期三則要多寫回 結算價
            row_for_settlement = write_row
            settlement_price = get_settlement_price(current_date)
            while sheet[f"E{row_for_settlement}"].value == week_code_for_0900:
                try:
                    C_settlement_price = settlement_price - int(sheet[f"H{row_for_settlement}"].value)
                    if C_settlement_price <= 0: C_settlement_price = 0
                except:
                    C_settlement_price = None

                try:
                    P_settlement_price = int(sheet[f"I{row_for_settlement}"].value) - settlement_price
                    if P_settlement_price <= 0: P_settlement_price = 0
                except:
                    P_settlement_price = None

                sheet[f"F{row_for_settlement}"].value = C_settlement_price
                sheet[f"F{row_for_settlement}"].alignment = alignment
                sheet[f"F{row_for_settlement}"].font = font
                sheet[f"F{row_for_settlement}"].number_format = "General"

                sheet[f"K{row_for_settlement}"].value = P_settlement_price
                sheet[f"K{row_for_settlement}"].alignment = alignment
                sheet[f"K{row_for_settlement}"].font = font
                sheet[f"K{row_for_settlement}"].number_format = "General"
            
                if current_date == sheet[f"A{row_for_settlement}"].value: # 加入下邊框
                    for col in range(1, 13):
                        sheet.cell(row=row_for_settlement, column=col).border = Border(bottom=Side(style='thin'))

                row_for_settlement = row_for_settlement - 1

            wb.save("Options.xlsx")

        write_row = write_row + 2

        print(f"".ljust(80, "="))
    
    print("已寫入完成 !")
    os.system("pause")